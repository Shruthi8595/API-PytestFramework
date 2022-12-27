import json
import jsonpath
import requests
import openpyxl

class Common:

    def __init__(self,FilePath,SheetName):
        global workbook
        global sheet
        workbook = openpyxl.load_workbook(FilePath)
        sheet = workbook[SheetName]

    def fetch_row_count(self):
        row = sheet.max_row
        return row

    def fetch_column_count(self):
        col = sheet.max_column
        return col

    def fetch_key_names(self):
        col = sheet.max_column
        list = []
        for i in range(1,col+1):
            cell = sheet.cell(row=1,column=i)
            list.insert(i-1,cell.value)
        return list

# If we want to update the particular row
    def update_requests_with_data(self,rowNumber,json_requests,keylist):
        col = sheet.max_column
        for i in range(1,col+1):
            cell = sheet.cell(row=rowNumber,column=i)
            json_requests[keylist[i-1]] = cell.value
            print(json_requests)
        return json_requests

