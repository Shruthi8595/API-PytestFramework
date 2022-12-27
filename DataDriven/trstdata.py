
import json
import openpyxl
import requests

def read_data():

    # Here we storing the URL and converting the data to json format
    API_URL = "http://thetestingworldapi.com/api/studentsDetails"
    open_file = open("C:/Users/91994/Desktop/Python_API_Data/Read_JsonFile.json")
    json_request = json.loads(open_file.read())

# by using Openpyxl we are fetching the data from excel
    open_workbook = openpyxl.load_workbook("C:\\Users\\91994\\Desktop\\Python_API_Data\\excel_data.xlsx")
    read_sheet = open_workbook['Sheet1']
    no_rows = read_sheet.max_row

    for i in range(2,no_rows+1):
        cell_first_name = read_sheet.cell(row=i, column=1)
        cell_middle_name = read_sheet.cell(row=i, column=2)
        cell_last_name = read_sheet.cell(row=i, column=3)
        cell_date_of_birth = read_sheet.cell(row=i, column=4)

 # assign the value to keys and storing it in json_request
        json_request['first_name'] = cell_first_name.value
        json_request['middle_name'] = cell_middle_name.value
        json_request['last_name'] = cell_last_name.value
        json_request['date_of_birth'] = cell_date_of_birth.value

        response = requests.post(API_URL, json_request)
        print(response.text)
        print(response.status_code)
